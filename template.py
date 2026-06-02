#!/usr/bin/env python3
import json
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Iterator

import openpyxl

EXCEL_PATH = Path("input.xlsx")
OUT_PATH = Path("ingest_output.json")


@dataclass
class Record:
    field_a: str
    field_b: float
    field_c: int


def iter_rows(path: Path) -> Iterator[Record]:
    wb = openpyxl.load_workbook(path, read_only=True)
    try:
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        _ = header_row  # skip header
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            yield Record(
                field_a=str(row[0]) if row[0] is not None else "",
                field_b=float(row[1]) if row[1] is not None else 0.0,
                field_c=int(row[2]) if row[2] is not None else 0,
            )
    finally:
        wb.close()


def main() -> int:
    records = list(iter_rows(EXCEL_PATH))
    data = [asdict(r) for r in records]
    OUT_PATH.write_text(json.dumps(data, indent=2, default=str), encoding="utf-8")
    print(f"Wrote {OUT_PATH}")
    return 0


if __name__ == "__main__":
    main()
