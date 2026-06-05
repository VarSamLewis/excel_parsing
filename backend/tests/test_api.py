"""Tests for the FastAPI routes — uses httpx TestClient.

These tests run in local-only mode.
"""

import json
import os
import subprocess
import unittest
from unittest.mock import patch
from fastapi.testclient import TestClient

# Ensure OpenAI settings are configured for tests
os.environ.setdefault("OPENAI_API_KEY", "test-key")

from backend.main import app

client = TestClient(app)


class TestHealth(unittest.TestCase):
    def test_health_returns_ok(self):
        response = client.get("/health")
        assert response.status_code == 200
        assert response.json() == {"status": "ok"}


class TestIngest(unittest.TestCase):
    """Test the /ingest endpoint with a mocked LLM call."""

    @patch("backend.main.generate_ingest_code")
    @patch("backend.main.infer_mapping")
    @patch("backend.main.validate_extraction")
    def test_ingest_basic(self, mock_validate, mock_infer, mock_codegen):
        from backend.models import (
            ExcelMapping,
            ColumnMapping,
            Transform,
            ValidationResult,
        )

        # Mock the LLM responses
        mock_infer.return_value = ExcelMapping(
            sheet_name="Sheet1",
            header_row=1,
            data_start_row=2,
            mappings=[
                ColumnMapping(
                    source_col="A",
                    target_field="company_name",
                    transform=Transform.STRIP,
                    notes="Column A is 'Company'",
                ),
            ],
            reasoning="Found headers at row 1",
        )
        mock_validate.return_value = ValidationResult(
            confidence=0.95,
            passed=True,
            summary="All good",
        )
        mock_codegen.return_value = "#!/usr/bin/env python3\nprint('ok')\n"

        # Create a minimal xlsx file
        import openpyxl
        import io

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "Company"
        ws["A2"] = "  Tesco PLC  "
        ws["A3"] = "  Sainsbury's  "

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        schema = {
            "name": "Test",
            "fields": [
                {
                    "name": "company_name",
                    "field_type": "string",
                    "description": "Company name",
                    "required": True,
                }
            ],
        }

        response = client.post(
            "/ingest",
            files={
                "file": (
                    "test.xlsx",
                    buf,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            params={
                "schema_name": "Test",
                "schema_json": json.dumps(schema),
            },
        )

        assert response.status_code == 200
        data = response.json()
        assert data["success"] is True
        assert data["row_count"] == 2
        assert len(data["data"]) == 2
        assert data["data"][0]["company_name"] == "Tesco PLC"
        assert data["data"][1]["company_name"] == "Sainsbury's"
        # Verify per-sheet structure
        assert "sheets" in data
        assert "sheet_names" in data
        assert len(data["sheet_names"]) >= 1
        assert data["schema_version"] >= 1


class TestVerify(unittest.TestCase):
    """Test verify endpoint with mocked LLM verifier."""

    @patch("backend.main.verify_generated_output")
    def test_verify_ingestion(self, mock_verify):
        mock_verify.return_value = "# Summary\n\nLooks good."
        response = client.post(
            "/verify-ingestion",
            params={
                "schema_json": "{}",
                "generated_code": "print('x')",
                "output_json": "[]",
                "run_id": "run_test",
            },
        )
        assert response.status_code == 200
        payload = response.json()
        assert "report_markdown" in payload

    @patch("backend.main.verify_generated_output")
    def test_verify_clean_skips_llm(self, mock_verify):
        rows = [{"company_name": "Acme", "amount": 10.5}]
        schema = {
            "name": "S",
            "fields": [
                {"name": "company_name", "field_type": "string", "required": True},
                {"name": "amount", "field_type": "number", "required": True},
            ],
        }
        response = client.post(
            "/verify-ingestion",
            params={
                "schema_json": json.dumps(schema),
                "generated_code": "print('x')",
                "output_json": json.dumps(rows),
                "run_id": "run_clean",
            },
        )
        assert response.status_code == 200
        payload = response.json()
        assert payload["llm_used"] is False
        assert "Numeric Diagnostics" in payload["report_markdown"]
        mock_verify.assert_not_called()

    @patch("backend.main.verify_generated_output")
    def test_verify_anomaly_calls_llm(self, mock_verify):
        mock_verify.return_value = "LLM notes"
        rows = [{"company_name": None, "amount": "bad"}]
        schema = {
            "name": "S",
            "fields": [
                {"name": "company_name", "field_type": "string", "required": True},
                {"name": "amount", "field_type": "number", "required": True},
            ],
        }
        response = client.post(
            "/verify-ingestion",
            params={
                "schema_json": json.dumps(schema),
                "generated_code": "print('x')",
                "output_json": json.dumps(rows),
                "use_llm": "true",
                "run_id": "run_bad",
            },
        )
        assert response.status_code == 200
        payload = response.json()
        assert payload["llm_used"] is True
        mock_verify.assert_called_once()

    @patch("backend.main.verify_generated_output")
    def test_verify_anomaly_without_llm(self, mock_verify):
        rows = [{"company_name": None, "amount": "bad"}]
        schema = {
            "name": "S",
            "fields": [
                {"name": "company_name", "field_type": "string", "required": True},
                {"name": "amount", "field_type": "number", "required": True},
            ],
        }
        response = client.post(
            "/verify-ingestion",
            params={
                "schema_json": json.dumps(schema),
                "generated_code": "print('x')",
                "output_json": json.dumps(rows),
                "use_llm": "false",
                "run_id": "run_bad_det",
            },
        )
        assert response.status_code == 200
        payload = response.json()
        assert payload["llm_used"] is False
        assert "Deterministic Findings" in payload["report_markdown"]
        mock_verify.assert_not_called()


@unittest.skipIf(os.environ.get("CI"), "CLI subprocess tests require local DB fixture")
class TestLogsCli(unittest.TestCase):
    """E2E tests for CLI logs commands."""

    def test_logs_commands(self):
        proc_runs = subprocess.run(
            ["python3", "cli/excel_ingest_cli.py", "logs", "runs", "--limit", "5"],
            check=True,
            capture_output=True,
            text=True,
        )
        assert proc_runs.returncode == 0

        proc_run = subprocess.run(
            [
                "python3",
                "cli/excel_ingest_cli.py",
                "logs",
                "run",
                "--run-id",
                "run_bad",
            ],
            check=True,
            capture_output=True,
            text=True,
        )
        assert proc_run.returncode == 0

        proc_usage = subprocess.run(
            [
                "python3",
                "cli/excel_ingest_cli.py",
                "logs",
                "usage",
                "--since-hours",
                "24",
            ],
            check=True,
            capture_output=True,
            text=True,
        )
        assert proc_usage.returncode == 0

        proc_runs_json = subprocess.run(
            [
                "python3",
                "cli/excel_ingest_cli.py",
                "logs",
                "runs",
                "--limit",
                "5",
                "--json",
            ],
            check=True,
            capture_output=True,
            text=True,
        )
        assert proc_runs_json.returncode == 0
        assert proc_runs_json.stdout.strip().startswith("[")
