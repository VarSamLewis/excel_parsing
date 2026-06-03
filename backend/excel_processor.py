"""Excel file processing: hashing, sheet inspection, and header/row sampling.

Uses openpyxl for .xlsx files. All functions operate on file bytes or a loaded
workbook — no LLM calls happen here.
"""

from __future__ import annotations

import hashlib
import io
import logging
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


def compute_file_hash(file_bytes: bytes) -> str:
    """Return the first 16 hex chars of the SHA-256 hash of the file bytes.

    This is the cache key prefix used throughout the system.
    """
    return hashlib.sha256(file_bytes).hexdigest()[:16]


def load_workbook_from_bytes(file_bytes: bytes) -> Workbook:
    """Load workbook from bytes; args: file_bytes (bytes); returns: Workbook."""
    return load_workbook(
        filename=io.BytesIO(file_bytes),
        read_only=True,
        data_only=True,
    )


def get_sheet_names(file_bytes: bytes) -> list[str]:
    """Return the list of sheet names in the workbook."""
    wb = load_workbook_from_bytes(file_bytes)
    names = wb.sheetnames
    wb.close()
    return names


def summarise_sheet(
    file_bytes: bytes,
    *,
    sheet_name: str | None = None,
    max_sample_values: int = 8,
    max_distinct_categorical: int = 15,
    max_cols: int = 52,
) -> dict[str, Any]:
    """Build a smart per-column summary of a sheet for the LLM.

    Instead of sending raw row samples, this analyses each column and produces
    a compact summary that gives the LLM better signal for mapping decisions.

    For each column, the summary includes:
        - header: the header cell value
        - column_letter: e.g. "A", "B"
        - non_empty_count: number of non-empty cells
        - first_values: first N non-empty values (preserving order)
        - last_values: last N non-empty values
        - distinct_values: unique values if the column looks categorical
        - dominant_type: the most common Python type name seen
        - type_inconsistencies: description of any type mismatches
    """
    from collections import Counter

    wb = load_workbook_from_bytes(file_bytes)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
            )
        ws: Worksheet = wb[sheet_name]
    else:
        ws = wb.active
        sheet_name = ws.title

    total_rows = ws.max_row or 0
    total_cols = ws.max_column or 0
    col_count = min(max_cols, total_cols) if total_cols else max_cols

    # Read all rows into columnar format
    # First pass: find header row and collect column data
    all_rows: list[list[Any]] = []
    for row in ws.iter_rows(max_col=col_count, values_only=True):
        all_rows.append(list(row))

    wb.close()

    if not all_rows:
        return {
            "sheet_name": sheet_name,
            "total_rows": 0,
            "total_cols": total_cols,
            "header_row": 0,
            "columns": [],
        }

    # Find header row (first row with mostly non-empty string cells)
    header_row_idx = 0
    for idx, row in enumerate(all_rows):
        non_empty = [c for c in row if c is not None]
        if non_empty and all(isinstance(c, str) for c in non_empty):
            header_row_idx = idx
            break
        if non_empty:
            header_row_idx = idx
            break

    headers = (
        all_rows[header_row_idx]
        if header_row_idx < len(all_rows)
        else [None] * col_count
    )
    data_rows = all_rows[header_row_idx + 1 :]

    column_summaries = []
    for col_idx in range(col_count):
        col_letter = get_column_letter(col_idx + 1)
        header_val = headers[col_idx] if col_idx < len(headers) else None

        # Collect all non-empty values for this column
        values = []
        for row in data_rows:
            if col_idx < len(row) and row[col_idx] is not None:
                values.append(row[col_idx])

        if not values:
            column_summaries.append(
                {
                    "column_letter": col_letter,
                    "header": str(header_val) if header_val is not None else None,
                    "non_empty_count": 0,
                    "first_values": [],
                    "last_values": [],
                    "dominant_type": None,
                    "type_inconsistencies": None,
                    "distinct_values": None,
                }
            )
            continue

        # First and last values
        first_vals = [str(v) for v in values[:max_sample_values]]
        last_vals = (
            [str(v) for v in values[-max_sample_values:]]
            if len(values) > max_sample_values
            else []
        )

        # Type analysis
        type_counts = Counter(type(v).__name__ for v in values)
        dominant_type = type_counts.most_common(1)[0][0]
        type_inconsistencies = None
        if len(type_counts) > 1:
            minority_types = {
                t: c for t, c in type_counts.items() if t != dominant_type
            }
            type_inconsistencies = (
                f"Mostly {dominant_type} ({type_counts[dominant_type]}/{len(values)}) "
                f"but also: {', '.join(f'{t} ({c})' for t, c in minority_types.items())}"
            )

        # Categorical detection: if distinct count is low relative to total
        distinct_raw = set(str(v) for v in values)
        distinct_values = None
        if len(distinct_raw) <= max_distinct_categorical:
            distinct_values = sorted(distinct_raw)

        column_summaries.append(
            {
                "column_letter": col_letter,
                "header": str(header_val) if header_val is not None else None,
                "non_empty_count": len(values),
                "first_values": first_vals,
                "last_values": last_vals,
                "dominant_type": dominant_type,
                "type_inconsistencies": type_inconsistencies,
                "distinct_values": distinct_values,
            }
        )

    return {
        "sheet_name": sheet_name,
        "total_rows": total_rows,
        "total_cols": total_cols,
        "header_row": header_row_idx + 1,  # 1-indexed
        "data_start_row": header_row_idx + 2,  # 1-indexed
        "columns": column_summaries,
    }


def read_data_rows(
    file_bytes: bytes,
    sheet_name: str,
    data_start_row: int,
    columns: list[str],
) -> list[dict[str, Any]]:
    """Read all data rows from the specified sheet starting at data_start_row.

    Args:
        file_bytes: Raw Excel file bytes.
        sheet_name: Which sheet to read.
        data_start_row: 1-indexed row where data begins.
        columns: List of column letters to read (e.g. ['A', 'B', 'D']).

    Returns:
        List of dicts mapping column letter → cell value for each row.
    """
    wb = load_workbook_from_bytes(file_bytes)
    ws: Worksheet = wb[sheet_name]

    rows: list[dict[str, Any]] = []
    row_num = data_start_row

    for row in ws.iter_rows(min_row=data_start_row, values_only=False):
        row_data: dict[str, Any] = {}
        for cell in row:
            col_letter = get_column_letter(cell.column)
            if col_letter in columns:
                row_data[col_letter] = cell.value
        # Skip completely empty rows
        if any(v is not None for v in row_data.values()):
            row_data["__row_num__"] = row_num
            rows.append(row_data)
        row_num += 1

    wb.close()
    return rows
