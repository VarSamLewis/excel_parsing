#!/usr/bin/env python3

"""Generate sample Excel files for local CLI/backend testing."""

from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet


def _write_sales_workbook(path: Path) -> None:
    """Create sales workbook; args: path (Path); returns: None."""
    wb: Workbook = Workbook()
    ws_candidate: Worksheet | None = wb.active
    if ws_candidate is None:
        raise RuntimeError("Workbook has no active worksheet")
    ws: Worksheet = ws_candidate
    ws.title = "Sales"
    ws.append(["company_name", "invoice_date", "amount", "paid"])
    ws.append(["Acme Ltd", date(2026, 1, 10), 1200.50, True])
    ws.append(["Globex", date(2026, 1, 14), 980.00, False])
    ws.append(["Initech", date(2026, 2, 2), 1550.75, True])
    wb.save(path)


def _write_people_workbook(path: Path) -> None:
    """Create people workbook; args: path (Path); returns: None."""
    wb: Workbook = Workbook()
    ws_candidate: Worksheet | None = wb.active
    if ws_candidate is None:
        raise RuntimeError("Workbook has no active worksheet")
    ws: Worksheet = ws_candidate
    ws.title = "Employees"
    ws.append(["name", "department", "start_date", "salary"])
    ws.append(["Alice Brown", "Operations", "2024-04-01", "55000"])
    ws.append(["Sam Carter", "Finance", "2023-09-15", "62000"])
    ws.append(["Ravi Patel", "Engineering", "2025-01-20", "78000"])
    wb.save(path)


def _write_multi_sheet_workbook(path: Path) -> None:
    """Create multi-sheet workbook; args: path (Path); returns: None."""
    wb: Workbook = Workbook()
    ws_orders_candidate: Worksheet | None = wb.active
    if ws_orders_candidate is None:
        raise RuntimeError("Workbook has no active worksheet")
    ws_orders: Worksheet = ws_orders_candidate
    ws_orders.title = "Orders"
    ws_orders.append(["order_id", "customer", "total"])
    ws_orders.append([1001, "Acme Ltd", 249.99])
    ws_orders.append([1002, "Globex", 499.50])

    ws_inventory: Worksheet = wb.create_sheet("Inventory")
    ws_inventory.append(["sku", "item_name", "qty", "warehouse"])
    ws_inventory.append(["SKU-001", "Widget A", 120, "North"])
    ws_inventory.append(["SKU-002", "Widget B", 80, "South"])
    wb.save(path)


def main() -> int:
    """Generate sample Excel set; args: none; returns: int."""
    out_dir: Path = Path("test_excels")
    out_dir.mkdir(parents=True, exist_ok=True)

    sales_path: Path = out_dir / "sales_sample.xlsx"
    people_path: Path = out_dir / "people_sample.xlsx"
    multi_path: Path = out_dir / "multi_sheet_sample.xlsx"

    _write_sales_workbook(sales_path)
    _write_people_workbook(people_path)
    _write_multi_sheet_workbook(multi_path)

    print(f"Wrote {sales_path}")
    print(f"Wrote {people_path}")
    print(f"Wrote {multi_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
